#!/usr/bin/env python3

import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

from datetime import datetime

currenttime = datetime.today().strftime('%Y-%m-%d_%H-%M')

route_list = [['DST','MASK','Gateway']]
rule_list = [['Name', 'Src', 'Dst', 'Service', 'Enabled', 'Action', 'NAT Policy', 'Description', 'RejectAction','Tag', 'Schedule', 'Log Enabled', 'Route Policy', 'Proxy']]
srv_list = [['Name','Description','Protocol','Port']]
host_list = [['Name','Description','Host IP Address','Network Address','Net Mask','Start Address','End Address']]
alias_list = [['Name','Description','Alias Name','Address','Interface']]
int_list = [['Name','Description','Interface Device Name','Enabled','Node Type','IP Address','Default Gateway','Net Mask','Secondary IP Addresses']]

tree = ET.parse(sys.argv[1])
root = tree.getroot()
    
# Excel Workbook
dest_filename = 'firewallconfig-' + currenttime + '.xlsx'
wb = Workbook()

# Worksheet : Route
ws1 = wb.active
ws1.title = "Routes"

row_count = 0
ws1.append(route_list[0])

# Route Parsing
for form in root.findall("./system-parameters/route/route-entry"):
    
    z = []
    for x in form:
        if x.tag == "dest-address":        
            z.append(x.text)
        elif x.tag == "mask":        
            z.append(x.text)
        elif x.tag == "gateway-ip":        
            z.append(x.text)
    ws1.append(z)

    
# Worksheet : Ruleset
ws2 = wb.create_sheet(title="Ruleset")

row_count = 0
ws2.append(rule_list[0])

# Rule Parsing
for form in root.findall("./abs-policy-list/abs-policy"):
    z = []
    rule_name=rule_src=rule_dst=rule_service=rule_enabled=rule_action=rule_nat=rule_desc=rule_reject=rule_tag=rule_schedule=rule_log=rule_route=rule_proxy = ""
    for x in form:
        if x.tag == "name":        
            rule_name = x.text
            pass
            
        if x.tag == "from-alias-list":
            #multi = []
            for a in x: 
                rule_src = a.text
            pass        
            
        if x.tag == "to-alias-list":        
            for a in x: 
                rule_dst = a.text
            pass
            
        if x.tag == "service":        
            rule_service = x.text
            pass
            
        if x.tag == "enabled":        
            rule_enabled = x.text
            pass
            
        if x.tag == "firewall":        
            rule_action = x.text
            pass
            
        if x.tag == "policy-nat":        
            rule_nat = x.text
            pass
            
        if x.tag == "description":        
            rule_desc = x.text
            pass
        
        if x.tag == "reject-action":        
            rule_reject = x.text
            pass
        
        if x.tag == "tag-list":        
            for a in x: 
                rule_tag = a.text
            pass
            
        if x.tag == "settings":        
            for a in x: 
                if a.tag == 'schedule':
                    rule_schedule = a.text
                if a.tag == 'log-enabled':
                    rule_log = a.text
                if a.tag == 'policy-routing':
                    rule_route = a.text
                if a.tag == 'proxy':
                    rule_proxy = a.text
            pass
            
    ws2.append([rule_name,rule_src,rule_dst,rule_service,rule_enabled,rule_action,rule_nat,rule_desc,rule_reject,rule_tag,rule_schedule,rule_log,rule_route,rule_proxy])

# Worksheet : Services
ws3 = wb.create_sheet(title="Services")

row_count = 0
ws3.append(srv_list[0])
    
for form in root.findall("./service-list/service"):
    z = []
    srv_name=srv_desc=srv_proto=srv_port = ""
    for x in form:
        if x.tag == "name":        
            srv_name = x.text
            pass
            
        if x.tag == "description":        
            srv_desc = x.text
            pass
            
        if x.tag == "service-item":
            
            for a in x: 

                for member in a:
                    #for item in member:
                    if member.tag == 'protocol':
                        srv_proto = member.text
                        if srv_proto == "0":
                            srv_proto = "HOPOPT"
                        elif srv_proto == "1":
                            srv_proto = "ICMP"
                        elif srv_proto == "2":
                            srv_proto = "IGMP"
                        elif srv_proto == "6":
                            srv_proto = "TCP"
                        elif srv_proto == "17":
                            srv_proto = "UDP"
                        elif srv_proto == "47":
                            srv_proto = "GRE"
                        elif srv_proto == "50":
                            srv_proto = "ESP"
                        elif srv_proto == "51":
                            srv_proto = "AH"
                        elif srv_proto == "89":
                            srv_proto = "OSPFIGP"

                    if member.tag == 'server-port':
                        srv_port = member.text
                ws3.append([srv_name,srv_desc,srv_proto,srv_port])
            pass
            
# Worksheet : Hosts
ws4 = wb.create_sheet(title="Address Groups")

row_count = 0
ws4.append(host_list[0])
    
for form in root.findall("./address-group-list/address-group"):
    z = []
    host_name=host_desc=host_ipaddr=host_netaddr=host_mask=host_start=host_end = ""
    for x in form:
        if x.tag == "name":        
            host_name = x.text
            pass
            
        if x.tag == "description":        
            host_desc = x.text
            pass
            
        if x.tag == "addr-group-member":
            for a in x: 
                for member in a:
                    if member.tag == 'host-ip-addr':
                        host_ipaddr = member.text
                    if member.tag == 'ip-network-addr':
                        host_netaddr = member.text
                    if member.tag == 'ip-mask':
                        host_mask = member.text
                    if member.tag == 'start-ip-addr':
                        host_start = member.text      
                    if member.tag == 'end-ip-addr':
                        host_end = member.text                              
            pass
    ws4.append([host_name,host_desc,host_ipaddr,host_netaddr,host_mask,host_start,host_end])

# Worksheet : Aliases
ws5 = wb.create_sheet(title="Group Aliases")

row_count = 0
ws5.append(alias_list[0])
    
for form in root.findall("./alias-list/alias"):
    z = []
    alias_name=alias_desc=alias_alias=alias_address=alias_int = ""
    for x in form:
        if x.tag == "name":        
            alias_name = x.text
            pass
            
        if x.tag == "description":        
            alias_desc = x.text
            pass
            
        if x.tag == "alias-member-list":
            for a in x: 
                for member in a:
                    if member.tag == 'aliasname':
                        alias_alias = member.text
                    if member.tag == 'address':
                        alias_address = member.text
                    if member.tag == 'interface':
                        alias_int = member.text
            pass
    ws5.append([alias_name,alias_desc,alias_alias,alias_address,alias_int])

# Worksheet : Interfaces
ws6 = wb.create_sheet(title="Interfaces")

row_count = 0
ws6.append(int_list[0])
    
for form in root.findall("./interface-list/interface"):
    z = []
    int_name=int_desc=int_dev=int_enabled=int_node=int_addr=int_gw=int_mask = ""
    secondaryip = ""
    
    for x in form:
        if x.tag == "name":        
            int_name = x.text
            pass
            
        if x.tag == "description":        
            int_desc = x.text
            pass
            
        if x.tag == "if-item-list":
            for a in x: 
                for physif in a:
                    for member in physif:
                        if member.tag == 'if-dev-name':
                            int_dev = member.text
                        if member.tag == 'enabled':
                            int_enabled = member.text
                        if member.tag == 'ip-node-type':
                            int_node = member.text
                        if member.tag == 'ip':
                            int_addr = member.text
                        if member.tag == 'default-gateway':
                            int_gw = member.text
                        if member.tag == 'netmask':
                            int_mask = member.text
                        if member.tag == 'secondary-ip-list':
                            for item in member:
                                for secip in item:
                                    secondaryip = secip.text + ", " + secondaryip
            pass
    ws6.append([int_name,int_desc,int_dev,int_enabled,int_node,int_addr,int_gw,int_mask,secondaryip])


wb.save(filename = dest_filename)
